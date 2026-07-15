#!/usr/bin/env python3
"""Parse session_content_with_order.md and generate Excel file."""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def parse_md_file(filepath):
    """Parse the markdown file and extract session data."""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Split by session headers
    session_pattern = r'### Session:\s*(\d+)\s*\n'
    parts = re.split(session_pattern, content)
    
    sessions = []
    # parts[0] is before the first session, then alternating: session_id, content
    for i in range(1, len(parts), 2):
        session_id = parts[i].strip()
        if i + 1 < len(parts):
            block = parts[i + 1]
        else:
            break
        
        # Parse markdown table rows (skip header and separator)
        rows = []
        lines = block.strip().split('\n')
        table_started = False
        for line in lines:
            line = line.strip()
            if not line.startswith('|'):
                if table_started:
                    break  # End of table
                continue
            
            cells = [c.strip() for c in line.split('|')]
            # Remove empty first/last from split
            cells = [c for c in cells if c != '']
            
            # Skip header row and separator
            if not cells:
                continue
            if cells[0] in ('时间', '---', '------'):
                table_started = True
                continue
            if all(c.startswith('-') for c in cells):
                continue
            if len(cells) < 4:
                continue
            
            table_started = True
            
            # Parse sender
            time_val = cells[0] if len(cells) > 0 else ''
            sender_raw = cells[1] if len(cells) > 1 else ''
            message = cells[2] if len(cells) > 2 else ''
            order_id = cells[3] if len(cells) > 3 else '-'
            order_status = cells[4] if len(cells) > 4 else '-'
            order_create_time = cells[5] if len(cells) > 5 else '-'
            
            # Clean sender
            if '客人' in sender_raw:
                sender = '客人'
            elif '房东' in sender_raw:
                sender = '房东'
            else:
                sender = sender_raw
            
            # Clean order fields
            if order_id == '-':
                order_id = ''
            if order_status == '-':
                order_status = ''
            if order_create_time == '-':
                order_create_time = ''
            
            # Clean message
            if message == '(空)':
                message = ''
            
            rows.append({
                'session_id': session_id,
                'time': time_val,
                'sender': sender,
                'message': message,
                'order_id': order_id,
                'order_status': order_status,
                'order_create_time': order_create_time,
            })
        
        if rows:
            sessions.append((session_id, rows))
    
    return sessions


def create_excel(sessions, output_path):
    """Create Excel file from parsed session data."""
    wb = Workbook()
    ws = wb.active
    ws.title = '会话数据'
    
    # Header style
    header_font = Font(name='Arial', bold=True, size=11)
    header_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Data style
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(vertical='top', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    # Headers
    headers = ['session_id', '时间', '发送方', '消息内容', '订单ID', '订单状态', '下单时间']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 22
    
    current_row = 2
    
    for sess_idx, (session_id, rows) in enumerate(sessions):
        # Add blank separator row between sessions (not before first)
        if sess_idx > 0:
            current_row += 1
        
        for row_data in rows:
            ws.cell(row=current_row, column=1, value=row_data['session_id']).font = data_font
            ws.cell(row=current_row, column=1).alignment = center_align
            
            ws.cell(row=current_row, column=2, value=row_data['time']).font = data_font
            ws.cell(row=current_row, column=2).alignment = center_align
            
            ws.cell(row=current_row, column=3, value=row_data['sender']).font = data_font
            ws.cell(row=current_row, column=3).alignment = center_align
            
            ws.cell(row=current_row, column=4, value=row_data['message']).font = data_font
            ws.cell(row=current_row, column=4).alignment = data_align
            
            ws.cell(row=current_row, column=5, value=row_data['order_id']).font = data_font
            ws.cell(row=current_row, column=5).alignment = center_align
            
            ws.cell(row=current_row, column=6, value=row_data['order_status']).font = data_font
            ws.cell(row=current_row, column=6).alignment = center_align
            
            ws.cell(row=current_row, column=7, value=row_data['order_create_time']).font = data_font
            ws.cell(row=current_row, column=7).alignment = center_align
            
            current_row += 1
    
    wb.save(output_path)
    print(f"Excel file saved: {output_path}")
    print(f"Total sessions: {len(sessions)}")
    total_rows = sum(len(rows) for _, rows in sessions)
    print(f"Total data rows: {total_rows}")
    print(f"Total Excel rows (with separators): {current_row - 1}")


if __name__ == '__main__':
    md_path = '/root/.openclaw/workspace/session_content_with_order.md'
    xlsx_path = '/root/.openclaw/workspace/session_content_with_order.xlsx'
    
    sessions = parse_md_file(md_path)
    if not sessions:
        print("ERROR: No sessions found in markdown file!")
    else:
        create_excel(sessions, xlsx_path)
        for sid, rows in sessions:
            print(f"  Session {sid}: {len(rows)} messages")
