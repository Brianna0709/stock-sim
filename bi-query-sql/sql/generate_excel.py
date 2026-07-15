#!/usr/bin/env python3
"""
Generate Excel from final JSONL data.
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INPUT = '/root/.openclaw/workspace/bi-query-sql/sql/sql2_final_complete.jsonl'
OUTPUT = '/root/.openclaw/workspace/session_checkin_fixed.xlsx'

# Load data
with open(INPUT) as f:
    rows = [json.loads(l) for l in f]

# Sort by session_id, then gmt_msg_gen
rows.sort(key=lambda r: (r['session_id'], r.get('gmt_msg_gen', '')))

# Order status mapping
STATUS_MAP = {
    '0': '未下单',
    '5': '待确认',
    '6': '已确认',
    '7': '已入住',
    '': '-',
    None: '-',
}

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = '会话内容'

# Define styles
header_font = Font(name='Arial', size=11, bold=True)
header_fill = PatternFill(start_color='B4D2FF', end_color='B4D2FF', fill_type='solid')
cell_font = Font(name='Arial', size=11)
wrap_alignment = Alignment(wrap_text=True, vertical='top')
header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

# Column widths
col_widths = {'A': 15, 'B': 22, 'C': 10, 'D': 60, 'E': 12}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Headers
headers = ['session_id', '时间', '发送方', '消息内容', '订单状态']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# Freeze first row
ws.freeze_panes = 'A2'

# Write data, grouped by session_id with empty rows between groups
current_row = 2
prev_session_id = None

for row_data in rows:
    session_id = row_data.get('session_id', '')
    
    # Add empty row between different sessions
    if prev_session_id is not None and session_id != prev_session_id:
        current_row += 1
    
    # Session ID
    cell = ws.cell(row=current_row, column=1, value=session_id)
    cell.font = cell_font
    cell.alignment = wrap_alignment
    
    # Time
    cell = ws.cell(row=current_row, column=2, value=row_data.get('gmt_msg_gen', ''))
    cell.font = cell_font
    cell.alignment = wrap_alignment
    
    # Sender
    is_host = row_data.get('is_from_phx_host', '')
    sender = '房东' if str(is_host) == '1' else '客人'
    cell = ws.cell(row=current_row, column=3, value=sender)
    cell.font = cell_font
    cell.alignment = wrap_alignment
    
    # Message content
    cell = ws.cell(row=current_row, column=4, value=row_data.get('valid_payload', ''))
    cell.font = cell_font
    cell.alignment = wrap_alignment
    
    # Order status
    os_val = row_data.get('order_status', '')
    if os_val is None:
        os_val = ''
    status_text = STATUS_MAP.get(os_val, os_val)
    cell = ws.cell(row=current_row, column=5, value=status_text)
    cell.font = cell_font
    cell.alignment = wrap_alignment
    
    prev_session_id = session_id
    current_row += 1

wb.save(OUTPUT)
print(f"Excel saved to: {OUTPUT}")
print(f"Total data rows: {len(rows)}")
print(f"Total sessions: {len(set(r['session_id'] for r in rows))}")
print(f"Total Excel rows (including separators): {current_row - 1}")
